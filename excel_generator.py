import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ─── ЦВЕТА SMILE EVENT CATERING ────────────────────────────────
COLOR_DARK = "1A1A1A"       # тёмный фон (как в PDF)
COLOR_GOLD = "C9975A"       # золотой акцент
COLOR_LIGHT = "F5F5F5"      # светлый фон строк
COLOR_WHITE = "FFFFFF"
COLOR_RED = "E74C3C"
COLOR_GREEN = "27AE60"

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=COLOR_WHITE):
    return Font(name='Arial', size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=COLOR_DARK):
    return Font(name='Arial', size=size, bold=bold, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def right():
    return Alignment(horizontal='right', vertical='center')

def generate_excel(
    event_data: dict,
    selected: list,
    staff: dict,
    food_total: float,
    staff_total: float,
    grand_total: float,
    version: str = "internal",
    discount_label: str = None,
    discount_amount: float = 0,
    final_total: float = None,
    logistics_result: dict = None,
) -> str:
    """
    Генерирует Excel смету для Smile Event Catering.
    Возвращает путь к файлу.
    """
    if final_total is None:
        final_total = grand_total

    guests = event_data.get('guests', 1)
    wb = openpyxl.Workbook()

    # ── Лист 1: СМЕТА ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Смета"

    # Ширина колонок
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16

    row = 1

    # ── ШАПКА ──────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "SMILE EVENT CATERING"
    ws[f'A{row}'].font = Font(name='Arial', size=16, bold=True, color=COLOR_GOLD)
    ws[f'A{row}'].fill = fill(COLOR_DARK)
    ws[f'A{row}'].alignment = center()
    ws.row_dimensions[row].height = 35
    row += 1

    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Больше, чем просто еда — мы создаём состояние праздника"
    ws[f'A{row}'].font = Font(name='Arial', size=9, italic=True, color=COLOR_GOLD)
    ws[f'A{row}'].fill = fill(COLOR_DARK)
    ws[f'A{row}'].alignment = center()
    ws.row_dimensions[row].height = 20
    row += 1

    row += 1  # пустая строка

    # ── ИНФОРМАЦИЯ О МЕРОПРИЯТИИ ────────────────────────────────
    info_fields = [
        ("СМЕТА №", event_data.get('number', '001')),
        ("КЛИЕНТ", event_data.get('client', '')),
        ("ФОРМАТ", event_data.get('format', '')),
        ("ГОСТЕЙ", str(event_data.get('guests', ''))),
        ("ДАТА", event_data.get('date', '')),
        ("МЕСТО", event_data.get('place', '')),
        ("ВРЕМЯ", event_data.get('time', '')),
    ]

    for label, value in info_fields:
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color="888888")
        ws[f'A{row}'].alignment = left()

        ws.merge_cells(f'C{row}:E{row}')
        ws[f'C{row}'] = value
        ws[f'C{row}'].font = body_font(size=10, bold=True)
        ws[f'C{row}'].alignment = left()
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # пустая строка

    # ── ИТОГОВЫЕ ПОКАЗАТЕЛИ ─────────────────────────────────────
    budget = event_data.get('budget', 0)
    target_weight = event_data.get('target_weight', 0)
    fact_weight = sum(item.get('weight_per_person', 0) for item in selected)

    summary_data = [
        ("БЮДЖЕТ", f"{budget:,.0f} руб.".replace(',', ' '), COLOR_DARK),
        ("СТОИМОСТЬ МЕНЮ", f"{food_total:,.0f} руб.".replace(',', ' '), COLOR_DARK),
        ("ИТОГО", f"{final_total:,.0f} руб.".replace(',', ' '), COLOR_GOLD),
        ("НА 1 ГОСТЯ", f"{final_total//guests:,.0f} руб.".replace(',', ' '), COLOR_DARK),
    ]

    for label, value, color in summary_data:
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color="888888")
        ws[f'A{row}'].alignment = left()

        ws.merge_cells(f'D{row}:E{row}')
        ws[f'D{row}'] = value
        ws[f'D{row}'].font = Font(name='Arial', size=11, bold=True, color=color)
        ws[f'D{row}'].alignment = right()
        ws.row_dimensions[row].height = 22
        row += 1

    # Превышение бюджета
    if final_total > budget:
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "⚠ ПРЕВЫШЕНИЕ"
        ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color=COLOR_RED)

        ws.merge_cells(f'D{row}:E{row}')
        ws[f'D{row}'] = f"+{final_total - budget:,.0f} руб.".replace(',', ' ')
        ws[f'D{row}'].font = Font(name='Arial', size=11, bold=True, color=COLOR_RED)
        ws[f'D{row}'].alignment = right()
        ws.row_dimensions[row].height = 20
        row += 1

    # Скидка/наценка
    if discount_label and discount_amount:
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = discount_label
        ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color=COLOR_GREEN)

        ws.merge_cells(f'D{row}:E{row}')
        ws[f'D{row}'] = f"{discount_amount:,.0f} руб.".replace(',', ' ')
        ws[f'D{row}'].font = Font(name='Arial', size=11, bold=True, color=COLOR_GREEN)
        ws[f'D{row}'].alignment = right()
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1  # пустая строка

    # ── ЗАГОЛОВОК ТАБЛИЦЫ МЕНЮ ─────────────────────────────────
    headers = ["№", "НАИМЕНОВАНИЕ И ОПИСАНИЕ", "КОЛ-ВО", "ЦЕНА", "СУММА"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font(size=9)
        cell.fill = fill(COLOR_DARK)
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    # ── ПОЗИЦИИ МЕНЮ ────────────────────────────────────────────
    current_cat = None
    item_num = 1

    for item in selected:
        cat = item.get('category', '')

        # Строка категории
        if cat != current_cat:
            current_cat = cat
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = cat
            cell.font = Font(name='Arial', size=9, bold=True, color=COLOR_WHITE)
            cell.fill = fill("555555")
            cell.alignment = left()
            cell.border = thin_border()
            ws.row_dimensions[row].height = 18
            row += 1

        # Строка блюда
        bg = COLOR_WHITE if item_num % 2 == 0 else COLOR_LIGHT

        cells_data = [
            (1, item_num, center()),
            (2, f"{item.get('name', '')}\n{item.get('desc', '')}", left()),
            (3, item.get('qty_label', ''), center()),
            (4, item.get('price_label', ''), right()),
            (5, item.get('total_label', ''), right()),
        ]

        for col_idx, value, alignment in cells_data:
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.font = body_font(size=9)
            cell.fill = fill(bg)
            cell.alignment = alignment
            cell.border = thin_border()

        ws.row_dimensions[row].height = 28
        row += 1
        item_num += 1

    row += 1

    # ── ПЕРСОНАЛ ────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "ПЕРСОНАЛ И СЕРВИС"
    ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color=COLOR_WHITE)
    ws[f'A{row}'].fill = fill("555555")
    ws[f'A{row}'].alignment = left()
    ws[f'A{row}'].border = thin_border()
    ws.row_dimensions[row].height = 18
    row += 1

    STAFF_PRICES = {"Менеджер": 12000, "Официант": 10000, "Повар": 10000, "Грузчик": 10000}
    staff_num = 1
    for role, qty in staff.items():
        price = STAFF_PRICES.get(role, 10000)
        total = qty * price
        bg = COLOR_WHITE if staff_num % 2 == 0 else COLOR_LIGHT

        cells_data = [
            (1, staff_num, center()),
            (2, role, left()),
            (3, f"{qty} чел.", center()),
            (4, f"{price:,} руб.".replace(',', ' '), right()),
            (5, f"{total:,} руб.".replace(',', ' '), right()),
        ]

        for col_idx, value, alignment in cells_data:
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.font = body_font(size=9)
            cell.fill = fill(bg)
            cell.alignment = alignment
            cell.border = thin_border()

        ws.row_dimensions[row].height = 22
        row += 1
        staff_num += 1

    row += 1

    # ── ЛОГИСТИКА (если есть) ───────────────────────────────────
    if logistics_result:
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "ЛОГИСТИКА"
        ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color=COLOR_WHITE)
        ws[f'A{row}'].fill = fill("555555")
        ws[f'A{row}'].alignment = left()
        ws[f'A{row}'].border = thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

        logistics_items = logistics_result.get('items', [])
        for i, l_item in enumerate(logistics_items, 1):
            bg = COLOR_WHITE if i % 2 == 0 else COLOR_LIGHT
            cells_data = [
                (1, i, center()),
                (2, l_item.get('name', ''), left()),
                (3, str(l_item.get('qty', '')), center()),
                (4, f"{l_item.get('price', 0):,} руб.".replace(',', ' '), right()),
                (5, f"{l_item.get('total', 0):,} руб.".replace(',', ' '), right()),
            ]
            for col_idx, value, alignment in cells_data:
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                cell.font = body_font(size=9)
                cell.fill = fill(bg)
                cell.alignment = alignment
                cell.border = thin_border()
            ws.row_dimensions[row].height = 22
            row += 1

        row += 1

    # ── ИТОГОВАЯ СТРОКА ─────────────────────────────────────────
    logistics_total = logistics_result['total_logistics'] if logistics_result else 0

    totals = [
        ("Стоимость меню:", food_total),
        ("Персонал:", staff_total),
    ]
    if logistics_total:
        totals.append(("Логистика:", logistics_total))
    totals.append(("ИТОГО:", final_total))

    for label, amount in totals:
        is_final = label == "ИТОГО:"
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(
            name='Arial', size=10 if is_final else 9,
            bold=is_final,
            color=COLOR_GOLD if is_final else COLOR_DARK
        )
        ws[f'A{row}'].alignment = right()
        ws[f'A{row}'].fill = fill(COLOR_DARK if is_final else COLOR_LIGHT)

        ws[f'E{row}'] = f"{amount:,.0f} руб.".replace(',', ' ')
        ws[f'E{row}'].font = Font(
            name='Arial', size=11 if is_final else 9,
            bold=True,
            color=COLOR_GOLD if is_final else COLOR_DARK
        )
        ws[f'E{row}'].alignment = right()
        ws[f'E{row}'].fill = fill(COLOR_DARK if is_final else COLOR_LIGHT)
        ws[f'E{row}'].border = thin_border()
        ws.row_dimensions[row].height = 22 if is_final else 18
        row += 1

    row += 2

    # ── УСЛОВИЯ ОПЛАТЫ ──────────────────────────────────────────
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "УСЛОВИЯ ОПЛАТЫ"
    ws[f'A{row}'].font = Font(name='Arial', size=9, bold=True, color=COLOR_WHITE)
    ws[f'A{row}'].fill = fill(COLOR_DARK)
    ws[f'A{row}'].alignment = left()
    ws.row_dimensions[row].height = 18
    row += 1

    conditions = [
        "• Предоплата 50% при подписании договора",
        "• Остаток 50% за 3 дня до мероприятия",
        "• Счёт действителен 5 рабочих дней",
        f"• Менеджер: {event_data.get('manager', 'Елена Смирнова')}",
    ]
    for condition in conditions:
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = condition
        ws[f'A{row}'].font = body_font(size=9)
        ws[f'A{row}'].alignment = left()
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1

    # ── ФУТЕР ───────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "Smile Event Catering  |  info@slcatering.ru  |  +7 (925) 605-38-50  |  slcatering.ru"
    ws[f'A{row}'].font = Font(name='Arial', size=8, color="888888")
    ws[f'A{row}'].fill = fill(COLOR_DARK)
    ws[f'A{row}'].alignment = center()
    ws.row_dimensions[row].height = 18

    # ── Лист 2: РЕДАКТИРУЕМАЯ СМЕТА ─────────────────────────────
    ws2 = wb.create_sheet("Редактировать")
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 16

    # Заголовок
    ws2['A1'] = "РЕДАКТИРУЕМАЯ СМЕТА"
    ws2['A1'].font = Font(name='Arial', size=14, bold=True, color=COLOR_GOLD)
    ws2['A1'].fill = fill(COLOR_DARK)
    ws2.merge_cells('A1:D1')
    ws2['A1'].alignment = center()
    ws2.row_dimensions[1].height = 30

    # Заголовки таблицы
    headers2 = ["НАИМЕНОВАНИЕ", "КОЛ-ВО (порц.)", "ЦЕНА (руб.)", "СУММА"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col_idx)
        cell.value = h
        cell.font = header_font(size=9)
        cell.fill = fill(COLOR_DARK)
        cell.alignment = center()
        cell.border = thin_border()
    ws2.row_dimensions[2].height = 20

    # Данные — редактируемые
    r = 3
    current_cat2 = None
    for item in selected:
        cat = item.get('category', '')
        if cat != current_cat2:
            current_cat2 = cat
            ws2.merge_cells(f'A{r}:D{r}')
            ws2[f'A{r}'] = cat
            ws2[f'A{r}'].font = Font(name='Arial', size=9, bold=True, color=COLOR_WHITE)
            ws2[f'A{r}'].fill = fill("555555")
            ws2[f'A{r}'].alignment = left()
            ws2.row_dimensions[r].height = 16
            r += 1

        qty = item.get('qty', 0)
        price = item.get('price', 0)

        ws2[f'A{r}'] = item.get('name', '')
        ws2[f'A{r}'].font = body_font(size=9)
        ws2[f'A{r}'].alignment = left()
        ws2[f'A{r}'].border = thin_border()

        ws2[f'B{r}'] = qty
        ws2[f'B{r}'].font = Font(name='Arial', size=9, color="0000FF")  # синий = редактируемое
        ws2[f'B{r}'].alignment = center()
        ws2[f'B{r}'].border = thin_border()

        ws2[f'C{r}'] = price
        ws2[f'C{r}'].font = Font(name='Arial', size=9, color="0000FF")  # синий = редактируемое
        ws2[f'C{r}'].alignment = right()
        ws2[f'C{r}'].border = thin_border()

        ws2[f'D{r}'] = f'=B{r}*C{r}'  # формула
        ws2[f'D{r}'].font = body_font(size=9)
        ws2[f'D{r}'].alignment = right()
        ws2[f'D{r}'].border = thin_border()

        ws2.row_dimensions[r].height = 20
        r += 1

    # Итого на листе 2
    r += 1
    ws2.merge_cells(f'A{r}:C{r}')
    ws2[f'A{r}'] = "ИТОГО МЕНЮ:"
    ws2[f'A{r}'].font = Font(name='Arial', size=10, bold=True)
    ws2[f'A{r}'].alignment = right()

    ws2[f'D{r}'] = f'=SUM(D3:D{r-2})'
    ws2[f'D{r}'].font = Font(name='Arial', size=11, bold=True, color=COLOR_GOLD)
    ws2[f'D{r}'].alignment = right()
    ws2[f'D{r}'].border = thin_border()
    ws2.row_dimensions[r].height = 22

    # Примечание
    r += 2
    ws2[f'A{r}'] = "💡 Синие ячейки можно редактировать — количество и цена пересчитаются автоматически"
    ws2[f'A{r}'].font = Font(name='Arial', size=8, color="888888", italic=True)
    ws2.merge_cells(f'A{r}:D{r}')

    # ── Сохранение ──────────────────────────────────────────────
    suffix = "внутренняя" if version == "internal" else "клиент"
    client_name = event_data.get('client', 'клиент').replace(' ', '_')
    filename = f"/tmp/Смета_{suffix}_{client_name}.xlsx"
    wb.save(filename)
    return filename
